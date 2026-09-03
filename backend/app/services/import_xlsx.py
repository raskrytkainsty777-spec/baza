"""Импорт стартовых данных заказчика из xlsx.

Два файла, которые он собрал руками до появления сервиса:

* «Риелторы» (Лист1): город · логин · коэф_комментов · подписчиков · имя_профиля · описание.
  Город «— город не указан» → донор без города, статус unclassified.
* «Посты» (лист «посты»): Комментарии · Просмотры · Комм. на 1000 просм. · Источник ·
  Суть оффера · Объект · Тип призыва · Кодовое слово · Дата · Текст поста · Ссылка.
  Разметка уже сделана (заказчиком, ИИ) — кладём как есть и помечаем ai_at, чтобы
  воркер разметки их не трогал.

Повторный запуск безопасен: аккаунты и посты — upsert по username / shortcode.

    python -m app.services.import_xlsx --realtors "Риелторы.xlsx" --posts "Посты.xlsx"
"""
import argparse
import asyncio
import datetime as dt
import re
import zoneinfo

import openpyxl
from sqlalchemy import select
from sqlalchemy.dialects.postgresql import insert

from ..db import SessionLocal
from ..models import IgAccount, LgCity, LgDonor, LgPost

MSK = zoneinfo.ZoneInfo("Europe/Moscow")
NO_CITY = {"", "—", "— город не указан", "не указан"}
EMPTY = {None, "", "—", "-"}
_SHORTCODE = re.compile(r"/(?:p|reel|reels)/([A-Za-z0-9_-]+)")


def _s(v) -> str | None:
    if v in EMPTY:
        return None
    return str(v).strip() or None


def _login(v) -> str:
    return str(v).strip().lstrip("@").rstrip("/").lower()


def _int(v) -> int | None:
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _rows(path: str, sheet: str | None = None) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    rows = [r for r in ws.iter_rows(values_only=True) if r and any(c is not None for c in r)]
    hdr = [str(c).strip() if c is not None else f"col{i}" for i, c in enumerate(rows[0])]
    return [dict(zip(hdr, r)) for r in rows[1:]]


async def _city_map(db, names: set[str]) -> dict[str, int]:
    """Города из файла: Москва — активна, остальные заводим выключенными,
    чтобы разметка заказчика не потерялась."""
    existing = {c.name: c.id for c in (await db.execute(select(LgCity))).scalars()}
    for name in sorted(names):
        if name not in existing:
            city = LgCity(name=name, is_active=(name == "Москва"))
            db.add(city)
            await db.flush()
            existing[name] = city.id
    return existing


async def import_realtors(path: str) -> dict:
    rows = _rows(path)
    rows = [r for r in rows if _s(r.get("логин"))]
    async with SessionLocal() as db:
        cities = await _city_map(db, {_s(r["город"]) for r in rows if _s(r.get("город")) not in NO_CITY})
        n_acc = n_donor = n_uncl = 0
        for r in rows:
            username = _login(r["логин"])
            city_name = _s(r.get("город"))
            city_id = cities.get(city_name) if city_name not in NO_CITY else None

            stmt = insert(IgAccount).values(
                username=username,
                full_name=_s(r.get("имя_профиля")),
                bio=_s(r.get("описание")),
                followers=_int(r.get("подписчиков")),
                roles="donor",
                city_id=city_id,
                city_source="manual" if city_id else None,
            ).on_conflict_do_update(
                index_elements=["username"],
                set_={
                    "full_name": _s(r.get("имя_профиля")),
                    "bio": _s(r.get("описание")),
                    "followers": _int(r.get("подписчиков")),
                    "roles": "donor",
                    "city_id": city_id,
                    "city_source": "manual" if city_id else None,
                },
            ).returning(IgAccount.id)
            account_id = (await db.execute(stmt)).scalar_one()
            n_acc += 1

            # донор: одна строка на (аккаунт, город); без города — unclassified
            q = select(LgDonor).where(LgDonor.account_id == account_id)
            q = q.where(LgDonor.city_id == city_id) if city_id else q.where(LgDonor.city_id.is_(None))
            donor = (await db.execute(q)).scalar_one_or_none()
            if donor is None:
                donor = LgDonor(
                    account_id=account_id, city_id=city_id,
                    status="new" if city_id else "unclassified",
                    intake_stage="posts", found_via="manual",
                    status_reason="импорт из xlsx заказчика",
                )
                db.add(donor)
                n_donor += 1
                if not city_id:
                    n_uncl += 1
        await db.commit()
    return {"accounts": n_acc, "donors_created": n_donor, "unclassified": n_uncl, "cities": len(cities)}


async def import_posts(path: str) -> dict:
    rows = _rows(path, "посты")
    rows = [r for r in rows if _s(r.get("Ссылка"))]
    now = dt.datetime.now(dt.timezone.utc)
    async with SessionLocal() as db:
        moscow = (await db.execute(select(LgCity).where(LgCity.name == "Москва"))).scalar_one_or_none()
        if moscow is None:
            moscow = LgCity(name="Москва", is_active=True)
            db.add(moscow)
            await db.flush()

        # источник → (account_id, donor_id). Источник без аккаунта заводим тут же.
        donors: dict[str, tuple[int, int]] = {}
        for username in {_login(r["Источник"]) for r in rows}:
            acc = (await db.execute(select(IgAccount).where(IgAccount.username == username))).scalar_one_or_none()
            if acc is None:
                acc = IgAccount(username=username, roles="donor", city_id=moscow.id, city_source="manual")
                db.add(acc)
                await db.flush()
            donor = (await db.execute(select(LgDonor).where(
                LgDonor.account_id == acc.id, LgDonor.city_id == moscow.id))).scalar_one_or_none()
            if donor is None:
                donor = LgDonor(account_id=acc.id, city_id=moscow.id, status="new",
                                intake_stage="posts", found_via="manual",
                                status_reason="импорт постов из xlsx")
                db.add(donor)
                await db.flush()
            donors[username] = (acc.id, donor.id)

        n_new = n_skip = 0
        for r in rows:
            m = _SHORTCODE.search(str(r["Ссылка"]))
            if not m:
                n_skip += 1
                continue
            account_id, donor_id = donors[_login(r["Источник"])]
            published = r.get("Дата")
            if isinstance(published, dt.datetime) and published.tzinfo is None:
                published = published.replace(tzinfo=MSK)   # выгрузка parser.im — в московском времени
            category = _s(r.get("Объект"))
            comments = _int(r.get("Комментарии")) or 0
            stmt = insert(LgPost).values(
                shortcode=m.group(1),
                donor_id=donor_id, account_id=account_id,
                city_id=moscow.id, city_source="donor",
                url=str(r["Ссылка"]).strip(),
                caption=_s(r.get("Текст поста")),
                published_at=published,
                product_type="clips" if "/reel" in str(r["Ссылка"]) else None,
                views=_int(r.get("Просмотры")),
                comments_count=comments, comments_count_prev=comments,
                offer=_s(r.get("Суть оффера")),
                category=category,
                cta_type=_s(r.get("Тип призыва")),
                code_word=_s(r.get("Кодовое слово")),
                # «продающий» в файле заказчика = есть объект продажи
                is_selling=category is not None,
                ai_at=now,
                monitor_status="active",
                collected_comments=0,
            ).on_conflict_do_nothing(index_elements=["shortcode"])
            res = await db.execute(stmt)
            if res.rowcount:
                n_new += 1
            else:
                n_skip += 1

        # у доноров с постами разовая петля по постам уже пройдена
        for _, donor_id in donors.values():
            d = await db.get(LgDonor, donor_id)
            if d and d.status == "new":
                d.status = "monitored"
                d.intake_stage = "done"
                d.status_changed_at = now
                d.status_reason = "посты за 45 дней импортированы из xlsx, разметка заказчика"
        await db.commit()
    return {"posts_new": n_new, "posts_skipped": n_skip, "sources": len(donors)}


async def main():
    ap = argparse.ArgumentParser(description="Импорт xlsx заказчика в базу")
    ap.add_argument("--realtors")
    ap.add_argument("--posts")
    args = ap.parse_args()
    if args.realtors:
        print("риелторы:", await import_realtors(args.realtors))
    if args.posts:
        print("посты:", await import_posts(args.posts))


if __name__ == "__main__":
    asyncio.run(main())
